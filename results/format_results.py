import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


jira_code = "CHEV_747"
related = pd.read_csv(f"/Users/expcaba/Downloads/relacionados_por_prueba_{jira_code}.csv")
generated = pd.read_csv(f"/Users/expcaba/Downloads/generados_por_prueba_{jira_code}.csv")

name_case = "{jira_code}_prueba_{i}_indice_{indice}.xlsx"
#print(related.columns)

pruebas_raw = related["prueba"]
pruebas_unique = []
for prueba in pruebas_raw:
    if prueba not in pruebas_unique:
        pruebas_unique.append(prueba)
#print(pruebas_unique)
i = 1
indice = 23
for prueba in pruebas_unique:
    related_prueba_perfect = related[(related["prueba"] == prueba) & (related["validacion"] == "Perfect")].reset_index(drop = True)[["prueba","validacion","gherkin_asociado"]]
    #print(related_prueba_perfect)
    generated_prueba_from_related = generated[(generated["prueba_objetivo"] == prueba) & (generated["tipo_generacion"] == "semi de la misma prueba")][["prueba_objetivo","gherkin_sugerido"]].reset_index(drop = True).rename(columns = {
        "prueba_objetivo":"prueba",
        "gherkin_sugerido":"Gherkin Generado a partir de similares"
    })
    generated_prueba_by_request = generated[(generated["prueba_objetivo"] == prueba) & (generated["tipo_generacion"] == "semi y perfect de otras pruebas")][["prueba_objetivo","gherkin_sugerido"]].reset_index(drop = True).rename(columns = {
        "prueba_objetivo":"prueba",
        "gherkin_sugerido":"Gherkin Generado a partir de solo requerimiento"
    })
    result_prueba = related_prueba_perfect.merge(generated_prueba_from_related, on="prueba", how="outer").merge(generated_prueba_by_request, on="prueba", how="outer")
    result_prueba = result_prueba.rename(columns = {
        "prueba":"Prueba",
        "gherkin_asociado":"Gherkin Match"
    })
    result_prueba["Gherkin Señalado"] = ''
    result_prueba["Se encontro componente funcional?"] = ""
    result_prueba["Se encontro match valido?"] = ""
    result_prueba["Evaluación Gherkin Generado a partir de similares"] = ""
    result_prueba["Evaluación Gherkin generado a partir de requerimiento"] = ""

    result_prueba = result_prueba[["Prueba","Gherkin Señalado","Gherkin Match","Gherkin Generado a partir de similares",
                                   "Gherkin Generado a partir de solo requerimiento","Se encontro componente funcional?",
                                   "Se encontro match valido?","Evaluación Gherkin Generado a partir de similares",
                                   "Evaluación Gherkin generado a partir de requerimiento"]]
    
    result_prueba.to_excel(name_case.format(jira_code = jira_code, i = i,indice = indice),index = False)

    # Load workbook and select sheet
    wb = load_workbook(name_case.format(jira_code = jira_code,i = i,indice = indice))
    ws = wb.active  # Get the first sheet

    # Apply top alignment to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top")

    # Save the formatted Excel file
    wb.save(name_case.format(jira_code = jira_code,i = i,indice = indice))

    i=i+1
    indice = indice+1

    #print(result_prueba.columns)


